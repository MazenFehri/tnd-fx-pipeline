"""
Write formatted Excel report (openpyxl + charts).
"""
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional

from openpyxl import Workbook
from openpyxl.chart import AreaChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent


def _auto_width(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        maxlen = 0
        for cell in col:
            try:
                v = len(str(cell.value)) if cell.value is not None else 0
                maxlen = max(maxlen, v)
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(maxlen + 2, 50)


def write_excel_report(
    conn: sqlite3.Connection,
    prediction_dict: Dict[str, Any],
    output_path: Optional[Path] = None,
) -> Path:
    """Three sheets: Daily prediction, History 90d, Rolling weights."""
    if output_path is None:
        d = prediction_dict.get("date") or datetime.utcnow().strftime("%Y-%m-%d")
        if isinstance(d, str) and "T" in d:
            d = d[:10]
        output_path = ROOT / "reports" / f"tnd_report_{d}.xlsx"
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # --- Sheet 1 ---
    ws1 = wb.active
    ws1.title = "Daily prediction"
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")
    alt_fill = PatternFill("solid", fgColor="E7EEF7")

    pred_date = prediction_dict.get("date", "")
    ws1["A1"] = "USD/TND FX Model — Daily"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1["A2"] = f"Date: {pred_date}"
    ws1["A4"] = "Metric"
    ws1["B4"] = "Value"
    ws1["C4"] = "Notes"
    for c in ("A4", "B4", "C4"):
        ws1[c].fill = hdr_fill
        ws1[c].font = hdr_font
        ws1[c].alignment = Alignment(horizontal="center")

    def fmt(v):
        if v is None:
            return "N/A"
        if isinstance(v, float):
            return round(v, 6)
        return v

    iv1 = prediction_dict.get("intrinsic_v1")
    iv2 = prediction_dict.get("intrinsic_v2")
    pf = prediction_dict.get("prev_fix")
    pct = prediction_dict.get("basket_ret_pct")
    rows = [
        ("Predicted intrinsic (basket only)", fmt(iv1), "intrinsic_v1"),
        ("Predicted intrinsic (full model)", fmt(iv2), "intrinsic_v2 = v1 + KF spread"),
        ("Previous BCT fixing", fmt(pf), "Level before FX move"),
        ("Estimated change (%)", fmt(pct), "Approx. from basket log-return"),
        ("EUR weight", fmt(prediction_dict.get("w_eurusd")), "90d rolling OLS"),
        ("GBP weight", fmt(prediction_dict.get("w_gbpusd")), ""),
        ("JPY weight", fmt(prediction_dict.get("w_usdjpy")), ""),
        ("Kalman spread", fmt(prediction_dict.get("kf_spread")), ""),
        ("R^2 (full-sample OLS)", fmt(prediction_dict.get("r_squared")), "ret_Fix on FX returns"),
    ]
    for i, (label, val, note) in enumerate(rows, start=5):
        ws1[f"A{i}"] = label
        ws1[f"A{i}"].font = Font(bold=True)
        ws1[f"B{i}"] = val
        ws1[f"C{i}"] = note
        if (i - 5) % 2 == 0:
            for col in ("A", "B", "C"):
                ws1[f"{col}{i}"].fill = alt_fill

    _auto_width(ws1)

    # --- Sheet 2: History 90 ---
    ws2 = wb.create_sheet("History (90 days)")
    q = """
    SELECT p.date, f.fix_mid, f.ib_rate, p.intrinsic_v1, p.intrinsic_v2,
           f.ib_rate - f.fix_mid AS spread
    FROM predictions p
    LEFT JOIN fx_rates f ON f.date = p.date
    ORDER BY p.date DESC
    LIMIT 90
    """
    cur = conn.execute(q)
    hist = cur.fetchall()
    hist = list(reversed(hist))

    ws2.append(["Date", "BCT Fix", "IB Rate", "Intrinsic V1", "Intrinsic V2", "Spread"])
    for row in hist:
        ws2.append(list(row))

    if len(hist) >= 2:
        n = len(hist) + 1
        chart = LineChart()
        chart.title = "BCT Fix vs Intrinsic V2"
        chart.y_axis.title = "TND"
        chart.x_axis.title = "Date"
        cats = Reference(ws2, min_col=1, min_row=2, max_row=n - 1)
        v1 = Reference(ws2, min_col=2, min_row=1, max_row=n - 1)
        v2 = Reference(ws2, min_col=5, min_row=1, max_row=n - 1)
        chart.add_data(v1, titles_from_data=True)
        chart.add_data(v2, titles_from_data=True)
        chart.set_categories(cats)
        try:
            chart.series[0].graphicalProperties.line.solidFill = "1F4E79"
            chart.series[1].graphicalProperties.line.solidFill = "C65911"
            chart.series[1].graphicalProperties.line.dashStyle = "dash"
        except Exception:
            pass
        ws2.add_chart(chart, "H2")

    _auto_width(ws2)

    # --- Sheet 3: Rolling weights ---
    ws3 = wb.create_sheet("Rolling weights")
    q3 = """
    SELECT date, w_eurusd, w_gbpusd, w_usdjpy
    FROM predictions
    ORDER BY date DESC
    LIMIT 90
    """
    cur3 = conn.execute(q3)
    wrows = list(reversed(cur3.fetchall()))
    ws3.append(["Date", "w_EURUSD", "w_GBPUSD", "w_USDJPY"])
    for row in wrows:
        ws3.append(row)

    if len(wrows) >= 2:
        n3 = len(wrows) + 1
        ach = AreaChart()
        ach.title = "Rolling OLS weights (90d)"
        ach.grouping = "stacked"
        ach.y_axis.title = "Weight"
        cats3 = Reference(ws3, min_col=1, min_row=2, max_row=n3 - 1)
        v_eur = Reference(ws3, min_col=2, min_row=1, max_row=n3 - 1)
        ach.add_data(v_eur, titles_from_data=True)
        v_gbp = Reference(ws3, min_col=3, min_row=1, max_row=n3 - 1)
        ach.add_data(v_gbp, titles_from_data=True)
        v_jpy = Reference(ws3, min_col=4, min_row=1, max_row=n3 - 1)
        ach.add_data(v_jpy, titles_from_data=True)
        ach.set_categories(cats3)
        ws3.add_chart(ach, "F2")

    _auto_width(ws3)

    wb.save(output_path)
    return output_path
