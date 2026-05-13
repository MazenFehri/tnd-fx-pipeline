"""
Walk-forward backtest for the USD/TND intrinsic-value model.

For each date t in the evaluation window:
  1. Train the basket OLS and AR(1) Kalman on data strictly before t.
  2. Produce intrinsic_v2(t) using FX returns observed at t and the spread
     forward-propagated one step from the most recent Kalman state.
  3. Compare to the actual fix_mid(t).

Outputs:
  - Per-date dataframe of (date, fix_mid, v1, v2, naive_rw, err_v2, err_rw).
  - Aggregate metrics: MAE, RMSE, MAPE, directional accuracy, Diebold-Mariano
    p-value vs random walk, residual Ljung-Box, spread ADF/KPSS (best-effort).
  - JSON dump to reports/backtest_metrics.json.
  - 4th sheet appended to the latest tnd_report_*.xlsx if present.

Run:  python backtest.py [--window 90 --warmup 200 --db data/tnd.db]
"""
from __future__ import annotations

import argparse
import json
import math
import sqlite3
from pathlib import Path
from typing import Any, Dict, Optional, Tuple

import numpy as np
import pandas as pd

from clean_returns import load_and_clean
from model import fit_ols, rolling_weights, kalman_filter_spread

ROOT = Path(__file__).resolve().parent
DEFAULT_DB = ROOT / "data" / "tnd.db"
OUT_JSON = ROOT / "reports" / "backtest_metrics.json"


# ---------------------------------------------------------------------------
# Walk-forward
# ---------------------------------------------------------------------------

def walk_forward(df: pd.DataFrame, window: int = 90, warmup: int = 200) -> pd.DataFrame:
    """
    Expanding-window walk forward starting at index `warmup`.

    Returns a dataframe indexed by date with columns:
      fix_mid, intrinsic_v1, intrinsic_v2, naive_rw, err_v2, err_rw,
      ret_basket, kf_state.
    """
    n = len(df)
    if n <= warmup + 2:
        raise ValueError(f"Not enough rows for backtest: have {n}, need > {warmup + 2}")

    fix = df["Fix_Mid"].values.astype(float)
    r_eur = df["ret_EURUSD"].values.astype(float)
    r_gbp = df["ret_GBPUSD"].values.astype(float)
    r_jpy = df["ret_USDJPY"].values.astype(float)
    spread = df["spread_pub"].values.astype(float)
    dates = df["date"].values

    rows = []
    for t in range(warmup, n):
        train = df.iloc[:t]  # strictly before t
        if train["spread_pub"].dropna().shape[0] < 30:
            continue

        # Rolling 90d weights ending at t-1
        roll = rolling_weights(train, min(window, len(train) - 1))
        w_last = roll.iloc[-1]
        if pd.isna(w_last["w_EURUSD"]):
            continue

        # Kalman state through t-1.
        # use_mle=False here is a deliberate speed tradeoff: MLE refit per fold
        # would dominate runtime (≈12s × N folds). The OLS-seed gives the same
        # filtered series within a few bps for AR(1) with |φ| close to unity.
        kf = kalman_filter_spread(train["spread_pub"], use_mle=False)
        if kf.empty:
            continue
        # Forward-propagate one step using AR(1)(c, phi) refit on train spread
        s = train["spread_pub"].dropna().values.astype(float)
        if len(s) < 3:
            continue
        yt, yl = s[1:], s[:-1]
        X = np.column_stack([np.ones(len(yt)), yl])
        beta, *_ = np.linalg.lstsq(X, yt, rcond=None)
        c, phi = float(beta[0]), float(beta[1])
        kf_pred_t = float(c + phi * float(kf.iloc[-1]))

        # Basket return at t (uses log-returns observed at t)
        ret_t = (
            float(w_last["w_const"])
            + float(w_last["w_EURUSD"]) * r_eur[t]
            + float(w_last["w_GBPUSD"]) * r_gbp[t]
            + float(w_last["w_USDJPY"]) * r_jpy[t]
        )
        prev_fix = fix[t - 1]
        if not (np.isfinite(prev_fix) and prev_fix > 0):
            continue
        v1 = prev_fix * math.exp(ret_t)
        v2 = v1 + kf_pred_t

        actual = fix[t]
        naive = prev_fix  # random walk baseline

        rows.append({
            "date": pd.Timestamp(dates[t]).strftime("%Y-%m-%d"),
            "fix_mid": actual,
            "intrinsic_v1": v1,
            "intrinsic_v2": v2,
            "naive_rw": naive,
            "err_v2": actual - v2,
            "err_rw": actual - naive,
            "ret_basket": ret_t,
            "kf_state": kf_pred_t,
        })

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Statistical tests
# ---------------------------------------------------------------------------

def _norm_sf(z: float) -> float:
    """Two-sided normal survival fn — pure numpy."""
    return math.erfc(abs(z) / math.sqrt(2))


def diebold_mariano(e1: np.ndarray, e2: np.ndarray, h: int = 1) -> Dict[str, float]:
    """
    DM test for equal predictive accuracy (squared-error loss).
    H0: E[d_t] = 0 where d_t = e1_t^2 - e2_t^2.
    Two-sided p-value. h is forecast horizon (1 here).
    """
    e1 = np.asarray(e1, dtype=float)
    e2 = np.asarray(e2, dtype=float)
    mask = np.isfinite(e1) & np.isfinite(e2)
    e1, e2 = e1[mask], e2[mask]
    if len(e1) < 10:
        return {"DM_stat": float("nan"), "DM_pvalue": float("nan"), "n": int(len(e1))}
    d = e1 ** 2 - e2 ** 2
    n = len(d)
    dbar = float(np.mean(d))
    # Newey-West-style long-run variance with h-1 lags
    gamma0 = float(np.var(d, ddof=1))
    lrv = gamma0
    for k in range(1, h):
        cov = float(np.mean((d[k:] - dbar) * (d[:-k] - dbar)))
        lrv += 2 * (1 - k / h) * cov
    if lrv <= 0:
        return {"DM_stat": float("nan"), "DM_pvalue": float("nan"), "n": n}
    stat = dbar / math.sqrt(lrv / n)
    return {"DM_stat": float(stat), "DM_pvalue": float(_norm_sf(stat)), "n": n}


def ljung_box(x: np.ndarray, lags: int = 10) -> Dict[str, float]:
    """Ljung-Box Q-stat with chi-square asymptotic p-value (no scipy needed)."""
    x = np.asarray(x, dtype=float)
    x = x[np.isfinite(x)]
    n = len(x)
    if n < lags + 2:
        return {"Q": float("nan"), "pvalue": float("nan"), "lags": lags}
    x = x - x.mean()
    denom = np.dot(x, x)
    if denom == 0:
        return {"Q": float("nan"), "pvalue": float("nan"), "lags": lags}
    Q = 0.0
    for k in range(1, lags + 1):
        r_k = np.dot(x[:-k], x[k:]) / denom
        Q += r_k ** 2 / (n - k)
    Q *= n * (n + 2)
    # Survival fn of chi^2 with `lags` dof via gammaincc
    try:
        from math import gamma  # noqa
        # Use scipy if available for accuracy; otherwise approximate via Wilson-Hilferty.
        try:
            from scipy.stats import chi2
            p = float(chi2.sf(Q, lags))
        except Exception:
            # Wilson-Hilferty approximation: ((Q/k)^(1/3) - (1 - 2/(9k))) / sqrt(2/(9k)) ~ N(0,1)
            k = lags
            z = ((Q / k) ** (1.0 / 3.0) - (1.0 - 2.0 / (9.0 * k))) / math.sqrt(2.0 / (9.0 * k))
            p = _norm_sf(z) / 2  # one-sided upper tail
        return {"Q": float(Q), "pvalue": float(p), "lags": lags}
    except Exception:
        return {"Q": float(Q), "pvalue": float("nan"), "lags": lags}


def stationarity(spread: pd.Series) -> Dict[str, Any]:
    """ADF + KPSS on the spread series. Uses statsmodels if available."""
    out: Dict[str, Any] = {"ADF_stat": None, "ADF_pvalue": None,
                           "KPSS_stat": None, "KPSS_pvalue": None,
                           "engine": None}
    try:
        from statsmodels.tsa.stattools import adfuller, kpss
        s = spread.dropna().values
        if len(s) >= 20:
            adf = adfuller(s, autolag="AIC")
            out["ADF_stat"] = float(adf[0])
            out["ADF_pvalue"] = float(adf[1])
            try:
                kp = kpss(s, regression="c", nlags="auto")
                out["KPSS_stat"] = float(kp[0])
                out["KPSS_pvalue"] = float(kp[1])
            except Exception:
                pass
            out["engine"] = "statsmodels"
    except ImportError:
        out["engine"] = "skipped (install statsmodels for ADF/KPSS)"
    return out


# ---------------------------------------------------------------------------
# Metrics aggregation
# ---------------------------------------------------------------------------

def summarize(bt: pd.DataFrame) -> Dict[str, float]:
    if bt.empty:
        return {}
    err = bt["err_v2"].values.astype(float)
    rw = bt["err_rw"].values.astype(float)
    actual = bt["fix_mid"].values.astype(float)
    pred = bt["intrinsic_v2"].values.astype(float)

    mae = float(np.mean(np.abs(err)))
    rmse = float(np.sqrt(np.mean(err ** 2)))
    mape = float(np.mean(np.abs(err / actual))) * 100.0

    # Directional accuracy: sign of Δactual vs sign of Δpred
    d_actual = np.diff(actual)
    d_pred = np.diff(pred)
    mask = (d_actual != 0) & (d_pred != 0) & np.isfinite(d_actual) & np.isfinite(d_pred)
    da = float((np.sign(d_actual[mask]) == np.sign(d_pred[mask])).mean() * 100.0) if mask.any() else float("nan")

    # Out-of-sample R²
    ss_res = float(np.sum(err ** 2))
    ss_tot = float(np.sum((actual - np.mean(actual)) ** 2))
    r2_oos = 1.0 - ss_res / ss_tot if ss_tot > 0 else float("nan")

    # RW baseline
    mae_rw = float(np.mean(np.abs(rw)))
    rmse_rw = float(np.sqrt(np.mean(rw ** 2)))

    dm = diebold_mariano(rw, err)  # H0: equal accuracy; positive DM means v2 better
    lb = ljung_box(err, lags=10)

    return {
        "N": int(len(bt)),
        "MAE": mae,
        "RMSE": rmse,
        "MAPE_pct": mape,
        "Directional_Accuracy_pct": da,
        "OOS_R2": r2_oos,
        "MAE_random_walk": mae_rw,
        "RMSE_random_walk": rmse_rw,
        "DM_stat": dm["DM_stat"],
        "DM_pvalue_vs_RW": dm["DM_pvalue"],
        "LjungBox_Q_lag10": lb["Q"],
        "LjungBox_p_lag10": lb["pvalue"],
    }


# ---------------------------------------------------------------------------
# Excel sheet append (optional)
# ---------------------------------------------------------------------------

def append_to_excel(bt: pd.DataFrame, summary: Dict[str, Any], stat: Dict[str, Any]) -> Optional[Path]:
    reports_dir = ROOT / "reports"
    if not reports_dir.exists():
        return None
    candidates = sorted(reports_dir.glob("tnd_report_*.xlsx"))
    if not candidates:
        return None
    target = candidates[-1]

    try:
        from openpyxl import load_workbook
        wb = load_workbook(target)
        # Remove existing sheet if present (so re-runs overwrite cleanly)
        for name in ("Backtest", "Backtest Trace"):
            if name in wb.sheetnames:
                del wb[name]

        ws = wb.create_sheet("Backtest")
        ws.append(["Metric", "Value"])
        for k, v in summary.items():
            ws.append([k, v])
        ws.append([])
        ws.append(["Spread stationarity", ""])
        for k, v in stat.items():
            ws.append([k, v])

        ws2 = wb.create_sheet("Backtest Trace")
        ws2.append(list(bt.columns))
        for _, row in bt.iterrows():
            ws2.append([row[c] for c in bt.columns])

        wb.save(target)
        return target
    except Exception as e:
        print(f"[append_to_excel] Skipped: {e}")
        return None


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", type=str, default=str(DEFAULT_DB))
    ap.add_argument("--window", type=int, default=90, help="rolling OLS window")
    ap.add_argument("--warmup", type=int, default=200, help="first eval index")
    args = ap.parse_args()

    conn = sqlite3.connect(args.db)
    df = load_and_clean(conn, lookback_days=10_000)
    conn.close()

    if df.empty:
        print("[backtest] No data — aborting.")
        return 1

    print(f"[backtest] {len(df)} rows · evaluating from index {args.warmup} · window {args.window}")
    bt = walk_forward(df, window=args.window, warmup=args.warmup)
    print(f"[backtest] produced {len(bt)} out-of-sample predictions")

    summary = summarize(bt)
    stat = stationarity(df["spread_pub"])

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps({
        "summary": summary,
        "spread_stationarity": stat,
        "warmup": args.warmup,
        "window": args.window,
    }, indent=2, default=str))
    print(f"[backtest] wrote {OUT_JSON}")

    trace_csv = ROOT / "reports" / "backtest_trace.csv"
    bt.to_csv(trace_csv, index=False)
    print(f"[backtest] wrote {trace_csv}")

    excel_target = append_to_excel(bt, summary, stat)
    if excel_target:
        print(f"[backtest] appended Backtest sheets to {excel_target}")

    # Console summary
    print("\n--- Summary ---")
    for k, v in summary.items():
        if isinstance(v, float):
            print(f"  {k:30s} {v:>14.6f}")
        else:
            print(f"  {k:30s} {v}")
    print("\nSpread stationarity:")
    for k, v in stat.items():
        print(f"  {k:20s} {v}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
